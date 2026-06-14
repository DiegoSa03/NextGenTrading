import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.worksheet.page import PageMargins

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "MICRO 029"

# Page Setup for Letter Size
ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

# Title
ws.merge_cells('A1:C1')
ws['A1'] = "MICRO 029"
ws['A1'].font = Font(bold=True, size=16)
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

# Headers
ws['A2'] = "FECHA"
ws['B2'] = "PRODUCIDO"
ws['C2'] = "OBSERVACIONES"

for col in ['A', 'B', 'C']:
    ws[col + '2'].font = Font(bold=True, size=12)
    ws[col + '2'].alignment = Alignment(horizontal="center", vertical="center")

# Column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 50

# Apply borders to all cells (title + header + 16 empty rows = 18 rows)
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

total_rows = 18  # 1 title + 1 header + 16 empty

for row in range(1, total_rows + 1):
    for col in range(1, 4):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border

# Row heights to fill the letter page
# Letter = 11 inches, margins 0.5+0.5 = 1 inch, usable = 10 inches = 720 points
# 18 rows -> 720 / 18 = 40 points each
for row in range(1, total_rows + 1):
    ws.row_dimensions[row].height = 40

wb.save("formato_micro_029.xlsx")
print("Excel created: 16 empty rows, fits letter page")
